#!/usr/bin/env python3

# Import modules
import sys
import os
import json
import pyzabbix
import getpass
import openpyxl
import time
import datetime

from datetime import datetime as dt
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.utils.cell import column_index_from_string
from openpyxl.styles import (PatternFill, Border, Side, Alignment, Font, GradientFill, colors, Color)

if '-G' not in sys.argv and '-H' not in sys.argv:
    print("Please enter hostgroup name as argument: <script.py> -G '<hostgroup>' or hosts divided by comma: <script.py> -H '<host>,<host>,<host>'")
    sys.exit()

if '-G' in sys.argv:
    group_index = sys.argv.index('-G')+1
    if group_index < len(sys.argv):
        host_group_name = sys.argv[group_index]
        
if '-H' in sys.argv:
    host_index = sys.argv.index('-H')+1
    if host_index < len(sys.argv):
        host_name = sys.argv[host_index]
        words = host_name.strip('"')
        words = host_name.strip("'")
        words = host_name.split(',')
        wordCount = len(words)
        hostCount = wordCount

# Input for Zabbix API, Username and Password, timeFrom, timeTill
apiPath = input(f"Set Zabbix API address please:\n")
#apiPath = "----------"

apiUsername = input("Type 1 Zabbix API username please: ")
#apiUsername = "-----------"

apiPassword = getpass.getpass("Type Zabbix API user password please: ")
#apiPassword = "-----------"

### IMPORTANT! ZABBIX ITEM FILTER, for example: 'icmpping[,5]' or 'icmpping'
itemFilter = {'key_': 'icmpping[,5]'}  

# Auth
try:
    zabbixApi = pyzabbix.ZabbixAPI(apiPath)
except Exception:
    print("Incorrect API path... Try again")	
try:
    zabbixApi.login(user=apiUsername, password=apiPassword)
except Exception:
    print("Incorrect login... Try again")
try:
    zabbixApi.auth
except Exception:
    print("Failed to authorize... Try again")	

# Frameworks for ICMP ping history getting 
now = time.localtime()

# Get the last day of last month by taking the first day of this month and subtracting 1 day.
tillTime = int(time.mktime((datetime.date(now.tm_year, now.tm_mon, 1) - datetime.timedelta(1)).timetuple()))
first = (datetime.date(now.tm_year, now.tm_mon, 1) - datetime.timedelta(1)).replace(day=1)
fromTime = int(time.mktime(first.timetuple()))

# Define folder name according to the previous month
folder_name = str(datetime.date(now.tm_year, now.tm_mon, 1) - datetime.timedelta(1))
folder_name = folder_name[:7]

# Excel files folder creation
## Get current directory
currentFolder = os.getcwd()

## Change working directory to previous month if it exists, else - make directory and change current directory
try: 
    os.chdir(f"./{folder_name}/")
except Exception:
    os.mkdir(folder_name, 0o700)
    os.chdir(f"./{folder_name}/")

# Get basic required hostgroup/host data
## If search hostgroups
if '-G' in sys.argv:
    groupHostInfo = zabbixApi.hostgroup.get(filter={"name": host_group_name}, output = ['groupid'], selectHosts=['hostid','host'])
    JSrequest0 = json.dumps(groupHostInfo)
    json_object0 = json.dumps(groupHostInfo, indent=4)
    data0 = json.loads(JSrequest0)
    gr_id_for_search = data0[0]['groupid']
    rawHostInfo = zabbixApi.host.get(groupids=gr_id_for_search, output = ['hostid','host','status'], selectInterfaces=['ip','port'])

## If search hosts    
if '-G' not in sys.argv and '-H' in sys.argv:
    rawHostInfo = zabbixApi.host.get(output = ['hostid','host','status'], selectInterfaces=['ip','port'])

JSrequest = json.dumps(rawHostInfo)

# Serializing json
json_object = json.dumps(rawHostInfo, indent=4)

data = json.loads(JSrequest)

if '-G' in sys.argv and '-H' not in sys.argv:
    hostCount = len(data)

# Create Workbook and active worksheet (SLA)
wb = Workbook()
ws = wb.active
ws = wb.worksheets[0]
ws.title = "SLA"

# Creation of the table title and headers
# ws - first sla sheet, ws2 - second problems sheet
ws['A1'].value = f"SLA report"
ws.append(["Host"] + ["IP"] + ["Port"] + ["avg ICMP ping (last month)"])

wb.create_sheet(f"Problems")
ws2 = wb[f"Problems"]
ws2.title = f"Problems"
ws2['A1'].value = f"Problems"

# Headers for the second sheet
ws2.append(['Hostname'] + ['Event ID'] + ['Problem'] + ['Started'] + ['Event ID'] + ['Resolved'])

# Prepare and combine data
## General information
for x in data:
    if '-H' in sys.argv and '-G' not in sys.argv:
        if str(x['host']) in host_name:
            hostid = int(x['hostid'])
            host = str(x['host'])
            ip = str(x['interfaces'][0]['ip']) 
            port = int(x['interfaces'][0]['port']) 

# SLA
## !!! HERE WE USE itemFilter VARIABLE, this data set at the head of script
            items = zabbixApi.item.get(filter=itemFilter, host=x['host'], output='extend', selectHosts=['host','name'])
            numList = list()
            sla = float()

            for item in items:
                values = zabbixApi.history.get(itemids=item['itemid'], output=['value'], time_from=fromTime, time_till = tillTime, history=item['value_type'])
                for historyValue in values:
                    val = int(historyValue['value'])
                    numList.append(val) 

            try:
                hostItemValueLength = int(len(numList))
            except Exception as e:
                print(f"Exception: {e}")
            try:
                hostItemValueSum = int(sum(numList))
            except Exception as e:
                print(f"Exception: {e}")
            try:
                sla = float((hostItemValueSum/hostItemValueLength) * 100)
            except Exception as e:
                print(f"Exception: {e}")

## Add SLA info into report
            ws.append([host] + [ip] + [port] + [sla])

# Problems data
## IF YOU WANT TO GET 'Unavailable by ICMP ping' PROBLEMS ONLY - ADD: prb = zabbixApi.event.get(filter={'name': 'Unavailable by ICMP ping'}, value=1, ... 
# Get created problems
            prb = zabbixApi.event.get(value=1, suppressed=0, hostids=hostid, selectHosts=['host'], select_acknowledges=['clock'], output=['r_eventid','value','suppressed','eventid','name','clock'], problem_time_from=fromTime, problem_time_till = tillTime)
            for pr in range(len(prb)):
                y = prb[pr]['r_eventid']
# Get resolved problems
                res = zabbixApi.event.get(eventid=y, value=0, suppressed=0, hostids=hostid, selectHosts=['host'], select_acknowledges=['clock'], output=['r_eventid','value','suppressed','eventid','name','clock'], problem_time_from=fromTime, problem_time_till = tillTime)  
# Add data to the problems worksheet
                ws2.append([prb[pr]['hosts'][0]['host']] + [int(prb[pr]['eventid'])] + [prb[pr]['name']] + [dt.fromtimestamp(int(prb[pr]['clock'])).strftime('%Y-%m-%d %H:%M:%S')] + [int(res[pr]['eventid'])] + [dt.fromtimestamp(int(res[pr]['clock'])).strftime('%Y-%m-%d %H:%M:%S')])

# If we search by hostgroups
    elif '-G' in sys.argv and '-H' not in sys.argv:           
        hostid = int(x['hostid'])
        host = str(x['host'])
        ip = str(x['interfaces'][0]['ip']) 
        port = int(x['interfaces'][0]['port']) 

# SLA
## !!! HERE WE USE itemFilter VARIABLE, this data set at the head of script
        items = zabbixApi.item.get(filter=itemFilter, host=x['host'], output='extend', selectHosts=['host','name'])
        numList = list()
        sla = float()

        for item in items:
            values = zabbixApi.history.get(itemids=item['itemid'], output=['value'], time_from=fromTime, time_till = tillTime, history=item['value_type'])
            for historyValue in values:
                val = int(historyValue['value'])
                numList.append(val) 

        try:
            hostItemValueLength = int(len(numList))
        except Exception as e:
            print(f"Exception: {e}")
        try:
            hostItemValueSum = int(sum(numList))
        except Exception as e:
            print(f"Exception: {e}")
        try:
            sla = float((hostItemValueSum/hostItemValueLength) * 100)
        except Exception as e:
            print(f"Exception: {e}")

# Add SLA info into report
        ws.append([host] + [ip] + [port] + [sla])

## Problems data
## IF YOU WANT TO GET 'Unavailable by ICMP ping' PROBLEMS ONLY - ADD: prb = zabbixApi.event.get(filter={'name': 'Unavailable by ICMP ping'}, value=1, ... 
# Get created problems
        prb = zabbixApi.event.get(value=1, suppressed=0, hostids=hostid, selectHosts=['host'], select_acknowledges=['clock'], output=['r_eventid','value','suppressed','eventid','name','clock'], problem_time_from=fromTime, problem_time_till = tillTime)
        for pr in range(len(prb)):
            y = prb[pr]['r_eventid']
# Get resolved problems
            res = zabbixApi.event.get(eventid=y, value=0, suppressed=0, hostids=hostid, selectHosts=['host'], select_acknowledges=['clock'], output=['r_eventid','value','suppressed','eventid','name','clock'], problem_time_from=fromTime, problem_time_till = tillTime)
# Add data to the problems worksheet
            ws2.append([prb[pr]['hosts'][0]['host']] + [int(prb[pr]['eventid'])] + [prb[pr]['name']] + [dt.fromtimestamp(int(prb[pr]['clock'])).strftime('%Y-%m-%d %H:%M:%S')] + [int(res[pr]['eventid'])] + [dt.fromtimestamp(int(res[pr]['clock'])).strftime('%Y-%m-%d %H:%M:%S')])

# Styles   
## Let's configurate the style of report
# First sheet
## Merge cells for title
maxCol = ws.max_column
minRow = ws.min_row
maxRow = ws.max_row
maxColLetter = get_column_letter(maxCol)
refCell = maxColLetter+str(maxRow)

ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=maxCol)

# Column width
for col in ws.iter_cols(min_row=2, max_row=hostCount+2, min_col=1, max_col=4):
     max_length = 25
     column = col[0].column_letter
     for cell in col:
         try: 
             if len(str(cell.value)) > max_length:
                 max_length = 45
         except:
             continue
     adjusted_width = (max_length)
     ws.column_dimensions[column].width = adjusted_width

# Alignment style
alignment_title=Alignment(
                horizontal='center',
                vertical='top',
                text_rotation=0,
                wrap_text=True,
                shrink_to_fit=False,
                indent=0
                   )

alignment=Alignment(
                horizontal='left',
                vertical='top',
                text_rotation=0,
                wrap_text=True,
                shrink_to_fit=False,
                indent=0
                   )

# Font styles
fontTitle = Font(
        name='Bahnschrift SemiBold',
        size=14,
        bold=False,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='FF000000'
            )

fontHeaders = Font(
        name='Bahnschrift',
        size=11,
        bold=False,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='FF000000'
            )

fontCells = Font(
        name='Bahnschrift Light',
        size=9,
        bold=False,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='FF000000'
            )

# Fill cells style
fillTitle = PatternFill(fill_type='solid', fgColor='FFCBA4')
fillHeaders = PatternFill(fill_type='solid', fgColor='96c8a2')
fillCells = PatternFill(fill_type='solid', fgColor='addfad')
fillAvail = PatternFill(fill_type='solid', fgColor='32CD32')
fillUnreach = PatternFill(fill_type='solid', fgColor='CD5C5C')

# Borders style
borderTitle = Side(border_style="thick", color="2f4f4f")
borderHeader = Side(border_style="medium", color="2f4f4f")
borderCells = Side(border_style="thin", color="2f4f4f")

# Title cell
ws['A1'].font = fontTitle
ws['A1'].fill = fillTitle
ws['A1'].alignment = alignment_title

# Headers
for cells in ws.iter_cols(min_row=2, max_row=2):
    for cell in cells:
        cell.font = fontHeaders
        cell.fill = fillHeaders
        cell.alignment = alignment
        cell.border = Border(top=borderHeader, bottom=borderHeader, left=borderHeader, right=borderHeader) 
        
# Fill cells skipping one
#for rows in range(minRow+2, maxRow, 1):
    #for cells in ws2.iter_cols(min_row=rows, max_row=rows):
        #for cell in cells:
            #cell.fill = fillCells

# Data cells
for cells in ws.iter_cols(min_row=3, max_row=maxRow):
    for cell in cells:
        styleObj = cell.coordinate
        ws[f'{styleObj}'].alignment = alignment
        ws[f'{styleObj}'].border = Border(top=borderCells, bottom=borderCells, left=borderCells, right=borderCells)
        if cell.value == "100":
            cell.fill = fillAvail
        elif cell.value == "uncomputed":
            cell.fill = fillUnreach

for i in range(1, ws.max_row):
    ws.row_dimensions[i].height = 30

# Second sheet
## Merge cells for title
maxCol2 = ws2.max_column
minRow2 = ws2.min_row
maxRow2 = ws2.max_row
maxColLetter2 = get_column_letter(maxCol2)
refCell2 = maxColLetter2+str(maxRow2)

ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=maxCol2)

# Styles
# Column width
for col in ws2.iter_cols(min_row=2, max_row=maxRow2, min_col=1, max_col=6):
     max_length = 25
     column = col[0].column_letter
     for cell in col:
         try: 
             if len(str(cell.value)) > max_length:
                 max_length = 45
         except:
             continue
     adjusted_width = (max_length)
     ws2.column_dimensions[column].width = adjusted_width

# Title cell
ws2['A1'].font = fontTitle
ws2['A1'].fill = fillTitle
ws2['A1'].alignment = alignment_title

# Headers
for cells in ws2.iter_cols(min_row=2, max_row=2):
    for cell in cells:
        cell.font = fontHeaders
        cell.fill = fillHeaders
        cell.alignment = alignment
        cell.border = Border(top=borderHeader, bottom=borderHeader, left=borderHeader, right=borderHeader) 

# Fill cells skipping one
#for rows in range(minRow2+2, maxRow2, 1):
    #for cells in ws2.iter_cols(min_row=rows, max_row=rows):
        #for cell in cells:
            #cell.fill = fillCells

# Data cells
for cells in ws2.iter_cols(min_row=3, max_row=maxRow2):
    for cell in cells:
        styleObj = cell.coordinate
        ws2[f'{styleObj}'].alignment = alignment
        ws2[f'{styleObj}'].border = Border(top=borderCells, bottom=borderCells, left=borderCells, right=borderCells)

for i in range(1, ws2.max_row):
    ws2.row_dimensions[i].height = 30

# IMPORTANT! Log out from Zabbix API
zabbixApi.user.logout()

# Save file (with hostgroup name if we search by hostgroup and strong name 'sla_report.xlsx' if we search by hosts)
if '-G' in sys.argv:
    wb.save(f"sla_{host_group_name}.xlsx")
elif '-H' in sys.argv:
    wb.save(f"sla_report.xlsx")
